'''hi there ,
analysing spreadsheets can be tedious and mundane worry not i have got your back 
using openpyexcel spreadsheets can be analysized in matter of seconds instead manual work which takes  days\

here we'll see the graphed stock price fluctuation with given rate of increase in a bullish market'''

from optparse import TitledHelpFormatter
from turtle import title
import openpyxl as xl
from openpyxl.chart import BarChart3D,Reference 

def stock_fluctuation(filename,rate):
    wb=xl.load_workbook(r"stock.xlsx") #loading an excel spreadsheet
    sheet=wb['Sheet1']# creating an object of sheet
    for row in range(2,sheet.max_row+1): #updating price to reduce by 10%
         cell=sheet.cell(row,3)
         corrected_price=cell.value*rate
         corrected_price_cell=sheet.cell(row,4)
         corrected_price_cell.value=corrected_price#saving th corrected pricing at new sheet
    
    chart =  BarChart3D()
    chart.title = "Stock fluctuations : Bullish Market"
    titles=Reference(
        sheet,
        min_row=1,
        max_row=sheet.max_row,
          min_col=2,
          max_col=2)
        
    values = Reference(sheet,  #taking values as instance of reference class
          min_row=1,
          max_row=sheet.max_row,
          min_col=2,
          max_col=4)
    
    chart.add_data(values) #passed values to chart
    chart.set_categories(titles)
    sheet.add_chart(chart,"E2")
  
    
    wb.save(filename)
#since a bullish market a +.245 or 24.5 % increase !!!
stock_fluctuation(r"stock.xlsx",1.245)






